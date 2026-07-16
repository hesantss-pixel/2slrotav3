"""
nf_parser.py — Parsers de notas fiscais para 2SL LOG
------------------------------------------------------
Extrai pedidos (nf, cidade, peso, itens, destinatario) a partir de:
  - XML NF-e (padrão SEFAZ 4.00)
  - Excel (.xlsx / .xls)
  - CSV (utf-8 / latin-1, separador , ou ;)
  - PDF (DANFE / notas em texto — extração por regex)

Retorna sempre uma lista de dicts no formato esperado pelo Firebase:
{
  "nf":            "301245",
  "dest":          "Cliente XYZ",
  "cidade":        "Guarulhos",
  "peso":          125.4,
  "itens":         3,
  "urgente":       False,
  "agendado":      False,
  "dataAgendada":  "",
  "horaEntrega":   "",
  "status":        "pendente",
  "lat":           -23.4566,
  "lng":           -46.5055,
  "origem":        "upload_xml" | "upload_xlsx" | "upload_csv" | "upload_pdf",
}
"""

from __future__ import annotations

import csv
import io
import logging
import re
import unicodedata
import xml.etree.ElementTree as ET
from typing import Iterable

logger = logging.getLogger(__name__)

DEPOT = {"lat": -23.4566, "lng": -46.5055}

CITY_COORDS: dict[str, dict] = {
    "Guarulhos":       {"lat": -23.4566, "lng": -46.5055},
    "Santos":          {"lat": -23.9617, "lng": -46.3322},
    "Guarujá":         {"lat": -23.9928, "lng": -46.2569},
    "São Paulo":       {"lat": -23.5614, "lng": -46.6559},
    "Praia Grande":    {"lat": -23.9864, "lng": -46.4119},
    "Sorocaba":        {"lat": -23.5015, "lng": -47.4581},
    "Itu":             {"lat": -23.2644, "lng": -47.2996},
    "Mogi das Cruzes": {"lat": -23.5234, "lng": -46.1845},
    "Santo André":     {"lat": -23.6738, "lng": -46.5438},
    "São Bernardo do Campo": {"lat": -23.6944, "lng": -46.5654},
    "Osasco":          {"lat": -23.5320, "lng": -46.7920},
    "Campinas":        {"lat": -22.9099, "lng": -47.0626},
}

CITY_NORMALIZED: dict[str, str] = {}
for nome in CITY_COORDS:
    _norm = unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode().lower()
    CITY_NORMALIZED[_norm] = nome


def _slug(s: str) -> str:
    if not s:
        return ""
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower().strip()


def _match_cidade(raw: str | None) -> str:
    """Devolve nome canônico de cidade conhecida ou 'Outras Cidades'."""
    if not raw:
        return "Outras Cidades"
    n = _slug(raw)
    if n in CITY_NORMALIZED:
        return CITY_NORMALIZED[n]
    for key, nome in CITY_NORMALIZED.items():
        if key and key in n:
            return nome
    return "Outras Cidades"


def _coords_para(cidade: str) -> dict:
    return CITY_COORDS.get(cidade, DEPOT)


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "")
    # Formatos comuns: "1.234,56" (BR)  |  "1234.56" (EN)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(re.sub(r"[^\d.\-]", "", s) or 0)
    except ValueError:
        return 0.0


def _to_int(v) -> int:
    try:
        return int(round(_to_float(v)))
    except Exception:
        return 0


def _build_pedido(*, nf: str, cidade: str, peso: float, itens: int,
                  dest: str = "", origem: str = "upload") -> dict:
    cid = _match_cidade(cidade)
    coord = _coords_para(cid)
    return {
        "nf":           str(nf).strip(),
        "dest":         (dest or "").strip(),
        "cidade":       cid,
        "peso":         round(_to_float(peso), 2),
        "itens":        max(1, _to_int(itens)),
        "urgente":      False,
        "agendado":     False,
        "dataAgendada": "",
        "horaEntrega":  "",
        "status":       "pendente",
        "lat":          coord["lat"],
        "lng":          coord["lng"],
        "origem":       origem,
    }


# ── XML NF-e ──────────────────────────────────────────────────────────────────

_NFE_NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}


def _findtext(elem, path: str, default: str = "") -> str:
    n = elem.find(path, _NFE_NS)
    return (n.text or default).strip() if n is not None and n.text else default


def parse_xml(content: bytes, filename: str = "") -> list[dict]:
    """Aceita um XML de NF-e (nfeProc, NFe ou lote com <nfeProc> repetido)."""
    text = content.decode("utf-8", errors="ignore")
    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        raise ValueError(f"XML inválido em {filename}: {e}") from e

    tag = root.tag.split("}")[-1]
    if tag in ("nfeProc", "NFe"):
        infs = root.findall(".//nfe:infNFe", _NFE_NS)
    elif tag in ("procEventoNFe", "envEvento"):
        return []
    else:
        infs = root.findall(".//nfe:infNFe", _NFE_NS)

    pedidos = []
    for inf in infs:
        ide  = inf.find("nfe:ide",  _NFE_NS)
        dest = inf.find("nfe:dest", _NFE_NS)
        transp = inf.find("nfe:transp", _NFE_NS)

        nf = _findtext(ide, "nfe:nNF") if ide is not None else ""
        if not nf:
            continue

        dest_nome = _findtext(dest, "nfe:xNome") if dest is not None else ""
        end = dest.find("nfe:enderDest", _NFE_NS) if dest is not None else None
        cidade = _findtext(end, "nfe:xMun") if end is not None else ""

        # Peso e volumes na tag <transp><vol>
        peso = 0.0
        volumes = 0
        if transp is not None:
            for vol in transp.findall("nfe:vol", _NFE_NS):
                peso    += _to_float(_findtext(vol, "nfe:pesoB") or _findtext(vol, "nfe:pesoL"))
                volumes += _to_int(_findtext(vol, "nfe:qVol"))

        # Fallback: soma dos itens em <det>
        itens = volumes or len(inf.findall("nfe:det", _NFE_NS)) or 1

        pedidos.append(_build_pedido(
            nf=nf, cidade=cidade, peso=peso, itens=itens,
            dest=dest_nome, origem="upload_xml",
        ))
    return pedidos


# ── CSV ───────────────────────────────────────────────────────────────────────

_HEADER_ALIASES = {
    "nf":     ("nf", "nota", "nnf", "numero_nf", "numero", "n_nf"),
    "dest":   ("dest", "destinatario", "cliente", "razao_social", "nome"),
    "cidade": ("cidade", "municipio", "xmun", "mun", "destino"),
    "peso":   ("peso", "pesob", "peso_bruto", "peso_kg", "kg"),
    "itens":  ("itens", "volumes", "qvol", "qtd", "quantidade", "vol"),
}


def _map_header(fieldnames: list[str]) -> dict:
    m: dict[str, str] = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        for original in fieldnames:
            key = _slug(original).replace(" ", "_").replace("-", "_")
            if key in aliases:
                m[canonical] = original
                break
    return m


def parse_csv(content: bytes, filename: str = "") -> list[dict]:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"Encoding não suportado em {filename}")

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return []
    m = _map_header(reader.fieldnames)
    if "nf" not in m:
        raise ValueError(
            f"CSV {filename}: coluna 'nf' não encontrada. "
            f"Colunas detectadas: {reader.fieldnames}"
        )

    pedidos = []
    for row in reader:
        nf = (row.get(m["nf"], "") or "").strip()
        if not nf:
            continue
        pedidos.append(_build_pedido(
            nf=nf,
            cidade=row.get(m.get("cidade", ""), "") if m.get("cidade") else "",
            peso=row.get(m.get("peso", ""), 0) if m.get("peso") else 0,
            itens=row.get(m.get("itens", ""), 1) if m.get("itens") else 1,
            dest=row.get(m.get("dest", ""), "") if m.get("dest") else "",
            origem="upload_csv",
        ))
    return pedidos


# ── Excel (.xlsx / .xls) ──────────────────────────────────────────────────────

def parse_xlsx(content: bytes, filename: str = "") -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl não instalado; adicione ao requirements.txt") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header = [str(h) if h is not None else "" for h in header]
    m = _map_header(header)
    if "nf" not in m:
        raise ValueError(
            f"Excel {filename}: coluna 'nf' não encontrada. Colunas: {header}"
        )
    idx = {k: header.index(v) for k, v in m.items()}

    pedidos = []
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        nf_v = row[idx["nf"]] if idx["nf"] < len(row) else None
        if nf_v is None or str(nf_v).strip() == "":
            continue
        pedidos.append(_build_pedido(
            nf=str(nf_v),
            cidade=row[idx["cidade"]] if "cidade" in idx and idx["cidade"] < len(row) else "",
            peso=row[idx["peso"]] if "peso" in idx and idx["peso"] < len(row) else 0,
            itens=row[idx["itens"]] if "itens" in idx and idx["itens"] < len(row) else 1,
            dest=row[idx["dest"]] if "dest" in idx and idx["dest"] < len(row) else "",
            origem="upload_xlsx",
        ))
    return pedidos


# ── PDF (DANFE / notas em texto) ──────────────────────────────────────────────

_RE_NF     = re.compile(r"(?:N[º°o.]?\s*(?:NF|NOTA)?\s*|N[º°o.]?\s*)(\d{4,9})", re.IGNORECASE)
_RE_PESO_B = re.compile(r"PESO\s+BRUTO[^0-9]{0,20}([\d.,]+)", re.IGNORECASE)
_RE_QVOL   = re.compile(r"(?:QTDE?\.?|QUANTIDADE)[\s\S]{0,10}?(?:VOL|VOLUME)[^0-9]{0,20}(\d{1,4})", re.IGNORECASE)
_RE_MUN    = re.compile(r"MUNIC[IÍ]PIO[^A-Z0-9]{0,15}([A-ZÁÉÍÓÚÃÕÂÊÔÇ ]{3,40})", re.IGNORECASE)


def parse_pdf(content: bytes, filename: str = "") -> list[dict]:
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError("pdfplumber não instalado; adicione ao requirements.txt") from e

    text_parts = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            try:
                text_parts.append(page.extract_text() or "")
            except Exception as e:
                logger.warning("PDF page extract falhou em %s: %s", filename, e)
    text = "\n".join(text_parts)
    if not text.strip():
        raise ValueError(f"PDF {filename}: nenhum texto extraível (possível scan/imagem).")

    nf_match     = _RE_NF.search(text)
    peso_match   = _RE_PESO_B.search(text)
    qvol_match   = _RE_QVOL.search(text)
    mun_match    = _RE_MUN.search(text)

    if not nf_match:
        raise ValueError(f"PDF {filename}: número da NF não localizado.")

    return [_build_pedido(
        nf     = nf_match.group(1),
        cidade = mun_match.group(1).title() if mun_match else "",
        peso   = peso_match.group(1) if peso_match else 0,
        itens  = qvol_match.group(1) if qvol_match else 1,
        dest   = "",
        origem = "upload_pdf",
    )]


# ── Dispatcher ────────────────────────────────────────────────────────────────

def parse_file(filename: str, content: bytes) -> list[dict]:
    """Escolhe o parser por extensão."""
    name = (filename or "").lower()
    if name.endswith(".xml"):
        return parse_xml(content, filename)
    if name.endswith(".csv") or name.endswith(".txt"):
        return parse_csv(content, filename)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(content, filename)
    if name.endswith(".xls"):
        raise ValueError(
            f"Formato .xls antigo não suportado — salve como .xlsx e envie novamente ({filename})."
        )
    if name.endswith(".pdf"):
        return parse_pdf(content, filename)
    raise ValueError(f"Formato não suportado: {filename}")


def parse_many(files: Iterable[tuple[str, bytes]]) -> dict:
    """Executa vários arquivos e agrega resultado + falhas por arquivo."""
    pedidos: list[dict] = []
    erros: list[dict] = []
    for name, blob in files:
        try:
            parsed = parse_file(name, blob)
            pedidos.extend(parsed)
        except Exception as e:
            logger.warning("Parse falhou em %s: %s", name, e)
            erros.append({"arquivo": name, "erro": str(e)})
    # Dedup por NF (última leitura ganha)
    dedup: dict[str, dict] = {}
    for p in pedidos:
        dedup[p["nf"]] = p
    return {"pedidos": list(dedup.values()), "erros": erros}
